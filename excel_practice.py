import xlwings as xw
import os

def convert_excel(file_path, status_label):
    try:
        status_label.config(text="Processing...", fg="blue")

        app = xw.App(visible=False)
        wb = app.books.open(file_path)

        for sheet in wb.sheets:
            if sheet.name != "indicators":
                continue

            used_range = sheet.used_range
            values = used_range.value
            formats = used_range.number_format

            # Make sure it's a 2D list even for single row/col
            if not isinstance(values[0], list):
                values = [values]
                formats = [formats]

            for i, row in enumerate(values):
                for j, val in enumerate(row):
                    fmt = formats[i][j] if isinstance(formats[i], list) else formats[i]
                    if val is None or isinstance(val, str) and str(val).startswith('='):
                        continue
                    if fmt and ('%' in fmt or '0%' in fmt):
                        continue
                    if isinstance(val, (int, float)):
                        used_range[i+1, j+1].value = val / 1.95583  # +1 because xlwings uses 1-based indexing

        base = os.path.basename(file_path)
        name, ext = os.path.splitext(base)
        output_path = os.path.join(os.getcwd(), f"{name}_EUR{ext}")
        wb.save(output_path)
        wb.close()
        app.quit()
        status_label.config(text="Conversion complete", fg="green")

    except Exception as e:
        status_label.config(text=f"Error: {str(e)}", fg="red")


import openpyxl
import xlwings as xw
import os
import copy
from openpyxl.cell.cell import MergedCell

def convert_excel(file_path, status_label):
    try:
        status_label.config(text="Processing...", fg="blue")

        # Step 1: Process all sheets except "indicators" using openpyxl
        wb_in = openpyxl.load_workbook(file_path)
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)  # Remove default sheet

        for sheet_name in wb_in.sheetnames:
            if sheet_name == "indicators":
                continue

            ws_in = wb_in[sheet_name]
            ws_out = wb_out.create_sheet(title=sheet_name)

            # Copy sheet-level properties
            ws_out.sheet_properties.tabColor = ws_in.sheet_properties.tabColor
            ws_out.freeze_panes = ws_in.freeze_panes
            if ws_in.auto_filter and ws_in.auto_filter.ref:
                ws_out.auto_filter.ref = ws_in.auto_filter.ref

            # Copy column dimensions
            for col_letter, dim in ws_in.column_dimensions.items():
                ws_out.column_dimensions[col_letter].width = dim.width
                ws_out.column_dimensions[col_letter].hidden = dim.hidden
                ws_out.column_dimensions[col_letter].outlineLevel = dim.outlineLevel

            # Copy row dimensions
            for row_idx, dim in ws_in.row_dimensions.items():
                ws_out.row_dimensions[row_idx].height = dim.height
                ws_out.row_dimensions[row_idx].hidden = dim.hidden
                ws_out.row_dimensions[row_idx].outlineLevel = dim.outlineLevel

            # Copy merged cells
            for merged_range in ws_in.merged_cells.ranges:
                ws_out.merge_cells(str(merged_range))

            # Copy cell values and styles
            for row in ws_in.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue

                    new_cell = ws_out.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value

                    # Convert numeric values
                    if cell.value is not None and isinstance(cell.value, (int, float)) and \
                       not (cell.number_format and ('%' in cell.number_format or '0%' in cell.number_format)):
                        new_cell.value = cell.value / 1.95583

                    # Copy styles
                    new_cell.font = copy.copy(cell.font)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.alignment = copy.copy(cell.alignment)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)

                    # Copy comment
                    if cell.comment:
                        new_cell.comment = copy.copy(cell.comment)

                    # Copy hyperlink
                    if cell.hyperlink:
                        new_cell.hyperlink = cell.hyperlink

        # Save temporary file without "indicators"
        temp_path = os.path.join(os.getcwd(), "no_indicators.xlsx")
        wb_out.save(temp_path)

        # Step 2: Process "indicators" using xlwings and append to final workbook
        app = xw.App(visible=False)
        wb_original = app.books.open(file_path)
        wb_temp = app.books.open(temp_path)

        try:
            sheet = wb_original.sheets['indicators']
            # Convert numeric values while preserving flags
            for row in sheet.used_range.rows:
                for cell in row:
                    val = cell.value
                    if val is None or (isinstance(val, str) and val.startswith('=')):
                        continue
                    if cell.number_format and ('%' in cell.number_format or '0%' in cell.number_format):
                        continue
                    if isinstance(val, (int, float)):
                        cell.value = val / 1.95583
        except Exception as e:
            sheet = None
            status_label.config(text=f"Error processing 'indicators': {str(e)}", fg="red")

        if sheet:
            # Copy "indicators" to end of processed workbook
            sheet.api.Copy(After=wb_temp.sheets[wb_temp.sheets.count - 1].api)

        # Save final file
        base = os.path.basename(file_path)
        name, ext = os.path.splitext(base)
        final_path = os.path.join(os.getcwd(), f"{name}_EUR{ext}")
        wb_temp.save(final_path)

        # Cleanup
        wb_original.close()
        wb_temp.close()
        app.quit()
        if os.path.exists(temp_path):
            os.remove(temp_path)

        status_label.config(text="Conversion complete", fg="green")

    except Exception as e:
        status_label.config(text=f"Error: {str(e)}", fg="red")





